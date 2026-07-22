import hashlib
import re
from datetime import date
from io import BytesIO
from typing import List
import os
import socket
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from database import (
    DB_PATH,
    add_custom_option,
    clear_all_data,
    count_import_batches,
    count_payroll_records,
    delete_custom_option,
    delete_import_batch,
    delete_payroll_records,
    delete_records_by_source,
    init_db,
    list_batches,
    list_custom_options,
    list_payroll_records,
    save_import_records,
    update_payroll_record,
    count_batches_by_source,
)
from hr_system_core import (
    BONUS_TYPE_OPTIONS,
    CASE_DELTA_FIELDS,
    CASE_FIELD_OPTIONS,
    CASE_OVERWRITE_FIELDS,
    CASE_TOTAL_COLS,
    COMPANY_OPTIONS,
    HQ_CASE_FIELD_OPTIONS,
    HEADQUARTERS_PROJECT,
    HR_COST_COLS,
    HR_ITEM_OPTIONS,
    HR_MANUAL_ITEM_OPTIONS,
    MONTHLY_TOTAL_COLS,
    PERSONAL_INCOME_COLS,
    PROJECT_OPTIONS,
    SITE_PROJECT_OPTIONS,
    SITE_COST_ITEM_OPTIONS,
    YEARLY_STAT_COLS,
    YEAR_OPTIONS,
    append_note_parts,
    build_case_total_frame,
    build_hr_cost_frame,
    build_personal_income_frame,
    build_monthly_total_frame,
    build_yearly_stat_frame,
    calc_hr_ratio,
    calc_request_pct,
    hq_revenue_base,
    delete_non_report_data,
    is_report_visible_source,
    build_hr_import_template_bytes,
    parse_hr_detail_workbook,
    parse_note_remark,
    parse_note_value,
    rebuild_note_display_fields,
    parse_hr_system_workbook,
    parse_note_number,
    roc_year_from_value,
)


def pick_manual_date(key_prefix: str, default: date | None = None) -> date:
    """可點選的日期欄，顯示格式 2024/04/04。"""
    value = st.date_input(
        "日期",
        value=default or date.today(),
        format="YYYY/MM/DD",
        key=key_prefix,
    )
    if isinstance(value, tuple):
        value = value[0] if value else date.today()
    return value


def format_currency_df(
    df: pd.DataFrame,
    cols: List[str],
    percent_cols: List[str] | None = None,
) -> pd.io.formats.style.Styler:
    percent_cols = percent_cols or []
    numeric_cols = [
        c
        for c in cols
        if c in df.columns and pd.api.types.is_numeric_dtype(pd.to_numeric(df[c], errors="coerce"))
    ]
    fmt = {c: "{:,.0f}" for c in numeric_cols if c not in percent_cols}
    fmt.update({c: "{:.2f}%" for c in numeric_cols if c in percent_cols})

    def style_rows(row: pd.Series) -> List[str]:
        values = [str(v).strip() for v in row.tolist()]
        is_total = "合計" in values
        if is_total:
            return ["background-color: #1d4ed8; color: #ffffff; font-weight: 700;" for _ in row]
        if isinstance(row.name, int) and row.name % 2 == 1:
            return ["background-color: rgba(37, 99, 235, 0.14);" for _ in row]
        return ["" for _ in row]

    styler = (
        df.style.format(fmt)
        .apply(style_rows, axis=1)
        .set_properties(subset=numeric_cols, **{"text-align": "right"})
        .set_properties(subset=[df.columns[0]], **{"text-align": "left"})
        .set_table_styles(
            [
                {
                    "selector": "th",
                    "props": [
                        ("background-color", "#1d4ed8"),
                        ("color", "#ffffff"),
                        ("font-weight", "700"),
                        ("text-align", "center"),
                        ("padding", "10px 8px"),
                        ("border", "1px solid #1e40af"),
                    ],
                },
                {
                    "selector": "td",
                    "props": [
                        ("padding", "8px"),
                        ("border", "1px solid rgba(120, 144, 180, 0.25)"),
                    ],
                },
            ],
            overwrite=False,
        )
    )
    return styler


def render_report_table(styled: pd.io.formats.style.Styler, max_height: int = 1200) -> None:
    """用 HTML 表格呈現，標題列可真正套用藍底（st.dataframe 做不到）。"""
    html = styled.hide(axis="index").to_html()
    st.markdown(
        f"""
        <div class="report-table-wrap" style="max-height:{max_height}px;">
        {html}
        </div>
        """,
        unsafe_allow_html=True,
    )


def to_excel_bytes(
    df: pd.DataFrame,
    sheet_name: str,
    numeric_cols: List[str],
    title: str,
    column_fill_map: dict[str, str] | None = None,
) -> bytes:
    column_fill_map = column_fill_map or {}
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.book[sheet_name]
        apply_sheet_style(ws, numeric_cols, title, column_fill_map)
    output.seek(0)
    return output.getvalue()


def apply_sheet_style(ws, numeric_cols: List[str], title: str, column_fill_map: dict[str, str]) -> None:
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    total_fill = PatternFill("solid", fgColor="E2F0D9")
    border = Border(
        left=Side(style="thin", color="666666"),
        right=Side(style="thin", color="666666"),
        top=Side(style="thin", color="666666"),
        bottom=Side(style="thin", color="666666"),
    )

    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=2, column=c)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for r in range(3, ws.max_row + 1):
        is_total = str(ws.cell(row=r, column=1).value).strip() == "合計"
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if is_total:
                cell.fill = total_fill
                cell.font = Font(bold=True)
            else:
                header = str(ws.cell(row=2, column=c).value or "")
                if header in column_fill_map:
                    cell.fill = PatternFill("solid", fgColor=column_fill_map[header])
            header = str(ws.cell(row=2, column=c).value or "")
            if header in numeric_cols:
                cell.number_format = '0.00"%"' if header == "比例" else "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    preferred_width = {
        "公司名": 20,
        "公司名稱": 20,
        "案場": 28,
        "姓名": 16,
        "分表": 14,
        "project_name": 28,
    }
    for c in range(1, ws.max_column + 1):
        col_letter = get_column_letter(c)
        header = str(ws.cell(row=2, column=c).value or "")
        if header in preferred_width:
            ws.column_dimensions[col_letter].width = preferred_width[header]
            continue
        if header in numeric_cols:
            ws.column_dimensions[col_letter].width = 13
            continue
        if c <= 2:
            ws.column_dimensions[col_letter].width = 20
            continue
        max_len = 0
        for r in range(1, ws.max_row + 1):
            value = ws.cell(row=r, column=c).value
            txt = "" if value is None else str(value)
            max_len = max(max_len, len(txt))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 28)

    ws.freeze_panes = "A3"


def to_excel_multi_sheets(
    sheet_data: dict[str, pd.DataFrame],
    numeric_cols: List[str],
    title_prefix: str,
    column_fill_map: dict[str, str] | None = None,
) -> bytes:
    column_fill_map = column_fill_map or {}
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in sheet_data.items():
            safe_name = str(sheet_name)[:31] if str(sheet_name).strip() else "Sheet"
            df.to_excel(writer, index=False, sheet_name=safe_name)
            ws = writer.book[safe_name]
            apply_sheet_style(ws, numeric_cols, f"{title_prefix}（{sheet_name}）", column_fill_map)
    output.seek(0)
    return output.getvalue()


def to_number(value: object) -> float:
    if pd.isna(value):
        return 0.0
    text = str(value).strip().replace(",", "").replace("%", "")
    if text in {"", "-", "nan", "None"}:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def detect_roc_year(text: str) -> int | None:
    match = re.search(r"(\d{3})年", text)
    if match:
        return int(match.group(1))
    return None


def clean_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


TOTAL_SUBSHEETS = ["勞健退", "薪資", "三節", "獎金", "餐費", "員工福利"]
TOTAL_MAIN_ITEM_COLS = ["勞健退", "薪資", "三節", "獎金", "餐費", "員工福利"]
TOTAL_MAIN_ALL_COLS = ["公司名", "案場", *TOTAL_MAIN_ITEM_COLS, "人事成本總計", "原始總計"]
PROJECT_SALES_COLS = ["案量", "銷售金額", "銷售佣金", "實收佣金", "廣告費", "服務費"]
PROJECT_HR_COLS = ["薪資", "勞健保", "退休金", "獎金"]
PROJECT_SUMMARY_COLS = ["公司名", "案名", *PROJECT_SALES_COLS, *PROJECT_HR_COLS, "人事成本合計"]
PROJECT_SUMMARY_ALL_COLS = PROJECT_SUMMARY_COLS
COMPANY_OPTIONS = ["得意佳", "匯鴻", "鴻源", "寶得", "得威", "賦鼎"]
PROJECT_OPTIONS = [
    "麗寶鐸藝",
    "天水一墅",
    "新潤世界都心",
    "名軒心城市",
    "麗寶之丘",
    "宏樸如嶼",
    "首學杭州",
    "首御臨沂",
    "商用不動產",
    "總公司",
]
INCOME_TYPE_OPTIONS = ["執行業務所得", "四倍獎金累計"]


def get_company_options() -> list:
    extra = [n for n in list_custom_options("company") if n not in COMPANY_OPTIONS]
    return [*COMPANY_OPTIONS, *extra]


def get_project_options() -> list:
    extra = [n for n in list_custom_options("project") if n not in PROJECT_OPTIONS]
    return [*PROJECT_OPTIONS, *extra]


def parse_manual_category(note: object) -> str:
    text = "" if pd.isna(note) else str(note)
    match = re.search(r"category:([^;]+)", text)
    if match:
        return match.group(1).strip()
    return ""


def parse_note_value(note: object, key: str) -> str:
    text = "" if pd.isna(note) else str(note)
    match = re.search(rf"{re.escape(key)}:([^;]+)", text)
    if match:
        return match.group(1).strip()
    return ""


def parse_note_number(note: object, key: str) -> float:
    return to_number(parse_note_value(note, key))


def append_note_parts(parts: list[str]) -> str:
    return ";".join([p for p in parts if str(p).strip()])


def get_local_ip() -> str:
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"


def calc_income_deductions(amount: float, income_type: str) -> tuple[float, float, float]:
    gross = float(amount or 0)
    if income_type == "執行業務所得":
        tax = round(gross * 0.10)
        health = 0.0
    else:
        tax = round(gross * 0.05)
        health = round(gross * 0.0211)
    net = round(gross - tax - health)
    return tax, health, net


def to_main_like_columns(df: pd.DataFrame, category: str) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=TOTAL_MAIN_ALL_COLS)

    out = pd.DataFrame()
    out["公司名"] = df["公司名"]
    out["案場"] = df["案場"]
    for item in TOTAL_MAIN_ITEM_COLS:
        out[item] = df["金額"] if category == item else 0.0
    out["人事成本總計"] = out[TOTAL_MAIN_ITEM_COLS].sum(axis=1)
    out["原始總計"] = out["人事成本總計"]
    return out


def build_main_total_frame(df_all: pd.DataFrame) -> pd.DataFrame:
    sr = df_all[df_all["source_type"] == "薪資占比"].copy()
    main_manual = df_all[df_all["source_type"] == "總表主表手動"].copy()
    main_adjust = df_all[df_all["source_type"] == "總表主表調整"].copy()

    rows = []
    for _, row in sr.iterrows():
        rows.append(
            {
                "公司名": str(row.get("company_name") or "").strip(),
                "案場": str(row.get("project_name") or "").strip(),
                "勞健退": 0.0,
                "薪資": float(row.get("salary") or 0),
                "三節": 0.0,
                "獎金": float(row.get("bonus") or 0),
                "餐費": 0.0,
                "員工福利": float(row.get("welfare") or 0),
                "原始總計": float(row.get("total_income") or 0),
            }
        )
    for _, row in main_manual.iterrows():
        salary = float(row.get("salary") or 0)
        bonus = float(row.get("bonus") or 0)
        welfare = float(row.get("welfare") or 0)
        original = float(row.get("total_income") or 0)
        labor = parse_note_number(row.get("note"), "勞健退")
        festival = parse_note_number(row.get("note"), "三節")
        meal = parse_note_number(row.get("note"), "餐費")
        if original <= 0:
            original = labor + salary + festival + bonus + meal + welfare
        rows.append(
            {
                "公司名": str(row.get("company_name") or "").strip(),
                "案場": str(row.get("project_name") or "").strip(),
                "勞健退": labor,
                "薪資": salary,
                "三節": festival,
                "獎金": bonus,
                "餐費": meal,
                "員工福利": welfare,
                "原始總計": original,
            }
        )
    for _, row in main_adjust.iterrows():
        rows.append(
            {
                "公司名": str(row.get("company_name") or "").strip(),
                "案場": str(row.get("project_name") or "").strip(),
                "勞健退": parse_note_number(row.get("note"), "勞健退"),
                "薪資": float(row.get("salary") or 0),
                "三節": parse_note_number(row.get("note"), "三節"),
                "獎金": float(row.get("bonus") or 0),
                "餐費": parse_note_number(row.get("note"), "餐費"),
                "員工福利": float(row.get("welfare") or 0),
                "原始總計": float(row.get("total_income") or 0),
            }
        )

    grouped = pd.DataFrame(rows)
    if grouped.empty:
        return pd.DataFrame(columns=TOTAL_MAIN_ALL_COLS)
    grouped = grouped.groupby(["公司名", "案場"], as_index=False)[[*TOTAL_MAIN_ITEM_COLS, "原始總計"]].sum()
    grouped["人事成本總計"] = grouped[TOTAL_MAIN_ITEM_COLS].sum(axis=1)
    return grouped[TOTAL_MAIN_ALL_COLS]


def build_total_subsheet_frames(df_all: pd.DataFrame) -> dict[str, pd.DataFrame]:
    bucket: dict[str, List[dict]] = {k: [] for k in TOTAL_SUBSHEETS}

    # 從薪資占比匯入資料拆成既有可對應分表
    sr = df_all[df_all["source_type"] == "薪資占比"].copy()
    for _, row in sr.iterrows():
        company = str(row.get("company_name") or "").strip()
        project = str(row.get("project_name") or "").strip()
        if company:
            if float(row.get("salary") or 0) > 0:
                bucket["薪資"].append({"公司名": company, "案場": project, "金額": float(row.get("salary") or 0), "備註": "來源: 薪資占比"})
            if float(row.get("bonus") or 0) > 0:
                bucket["獎金"].append({"公司名": company, "案場": project, "金額": float(row.get("bonus") or 0), "備註": "來源: 薪資占比"})
            if float(row.get("welfare") or 0) > 0:
                bucket["員工福利"].append({"公司名": company, "案場": project, "金額": float(row.get("welfare") or 0), "備註": "來源: 薪資占比"})

    # 手動輸入資料
    manual = df_all[
        df_all["source_type"].isin(["總表分表手動", "總表分表匯入"])
        | df_all["source_type"].fillna("").astype(str).str.startswith("總表分表調整_")
    ].copy()
    for _, row in manual.iterrows():
        category = parse_manual_category(row.get("note"))
        if category not in bucket:
            continue
        bucket[category].append(
            {
                "公司名": str(row.get("company_name") or "").strip(),
                "案場": str(row.get("project_name") or "").strip(),
                "金額": float(row.get("total_income") or 0),
                "備註": str(row.get("note") or "").replace(f"category:{category};", "").strip(),
            }
        )

    frames: dict[str, pd.DataFrame] = {}
    for category, rows in bucket.items():
        if rows:
            df = pd.DataFrame(rows)
            df = df.groupby(["公司名", "案場"], as_index=False)["金額"].sum()
            total_row = pd.DataFrame([{"公司名": "合計", "案場": "", "金額": df["金額"].sum()}])
            frames[category] = pd.concat([df, total_row], ignore_index=True)
        else:
            frames[category] = pd.DataFrame(columns=["公司名", "案場", "金額"])
    return frames


def subsheet_amount_map(df_all: pd.DataFrame) -> dict[tuple[str, str], dict[str, float]]:
    frames = build_total_subsheet_frames(df_all)
    out: dict[tuple[str, str], dict[str, float]] = {}
    for category, df in frames.items():
        if df.empty:
            continue
        for _, row in df.iterrows():
            company = str(row.get("公司名") or "").strip()
            project = str(row.get("案場") or "").strip()
            if not company or company == "合計" or not project:
                continue
            key = (company, project)
            out.setdefault(key, {})
            out[key][category] = out[key].get(category, 0.0) + float(row.get("金額") or 0)
    return out


def build_project_summary_frame(df_all: pd.DataFrame) -> pd.DataFrame:
    sources = df_all[df_all["source_type"].isin(["薪資占比", "總表主表手動", "總表主表調整"])].copy()
    sub_map = subsheet_amount_map(df_all)
    rows: list[dict] = []

    def append_row(company: str, project: str, note: object, salary: float, bonus: float, welfare: float) -> None:
        if not company or not project:
            return
        key = (company, project)
        sub = sub_map.get(key, {})
        hr_salary = sub.get("薪資", 0.0) or salary
        hr_labor = sub.get("勞健退", 0.0)
        hr_pension = 0.0
        hr_bonus = sub.get("獎金", 0.0) or bonus
        if hr_salary <= 0 and salary > 0:
            hr_salary = salary
        if hr_bonus <= 0 and bonus > 0:
            hr_bonus = bonus
        if hr_labor <= 0:
            hr_labor = parse_note_number(note, "勞健退")
        rows.append(
            {
                "公司名": company,
                "案名": project,
                "案量": parse_note_number(note, "全案總銷"),
                "銷售金額": parse_note_number(note, "簽約金額"),
                "銷售佣金": parse_note_number(note, "總銷1%"),
                "實收佣金": parse_note_number(note, "簽約金額1%"),
                "廣告費": parse_note_number(note, "營業收入"),
                "服務費": parse_note_number(note, "請款淨額"),
                "薪資": hr_salary,
                "勞健保": hr_labor,
                "退休金": hr_pension,
                "獎金": hr_bonus,
            }
        )

    for _, row in sources.iterrows():
        company = str(row.get("company_name") or "").strip()
        project = str(row.get("project_name") or "").strip()
        append_row(
            company,
            project,
            row.get("note"),
            float(row.get("salary") or 0),
            float(row.get("bonus") or 0),
            float(row.get("welfare") or 0),
        )

    if not rows:
        return pd.DataFrame(columns=PROJECT_SUMMARY_ALL_COLS)

    grouped = pd.DataFrame(rows).groupby(["公司名", "案名"], as_index=False).sum(numeric_only=True)
    grouped["人事成本合計"] = grouped[PROJECT_HR_COLS].sum(axis=1)
    return grouped[PROJECT_SUMMARY_ALL_COLS]


def parse_total_subsheet_workbook(file_path: str) -> List[dict]:
    xls = pd.ExcelFile(file_path)
    records: List[dict] = []

    for category in TOTAL_SUBSHEETS:
        if category not in xls.sheet_names:
            continue
        df = pd.read_excel(file_path, sheet_name=category, header=3)
        if df.empty or df.shape[1] < 3:
            continue

        col_company = df.columns[0]
        col_project = df.columns[1]
        month_cols = [c for c in df.columns[2:] if isinstance(c, str) and "/" in c]
        if not month_cols:
            month_cols = list(df.columns[2:])

        current_company = ""
        for _, row in df.iterrows():
            company_raw = clean_text(row.get(col_company))
            project = clean_text(row.get(col_project))
            if company_raw:
                current_company = company_raw.replace("\n", "")
            company = current_company
            if not company or not project:
                continue

            values = [to_number(row.get(c)) for c in month_cols]
            amount = float(sum(values))
            if amount <= 0:
                continue

            records.append(
                {
                    "sheet_name": "全案總表分表",
                    "employee_name": "未指定",
                    "company_name": company,
                    "project_name": project,
                    "roc_year": None,
                    "salary": 0.0,
                    "bonus": 0.0,
                    "welfare": 0.0,
                    "total_income": amount,
                    "note": f"category:{category};來源:薪資占比檔分表",
                }
            )
    return records


def upsert_adjustment_records(source_type: str, records: List[dict]) -> None:
    delete_records_by_source(source_type)
    if records:
        save_import_records(source_type, "inline_edit", records)


def parse_year_amounts(note: object) -> dict[str, float]:
    text = "" if pd.isna(note) else str(note)
    pairs = re.findall(r"(\d{2,4})\s*:\s*([-\d.]+)", text)
    out: dict[str, float] = {}
    for y, v in pairs:
        out[y] = float(pd.to_numeric(v, errors="coerce") or 0.0)
    return out


def parse_personal_income_detail_sheet(df: pd.DataFrame, sheet_name: str) -> List[dict]:
    records: List[dict] = []
    n_rows, n_cols = df.shape
    roc_year = detect_roc_year(sheet_name) or 114

    for i in range(n_rows):
        row = [clean_text(v) for v in df.iloc[i].tolist()]
        for start in range(max(0, n_cols - 2)):
            employee = row[start]
            if not employee:
                continue
            if start + 2 >= n_cols:
                continue
            if "案場" not in row[start + 1] or "金額" not in row[start + 2]:
                continue

            employee_name = employee.replace("*", "").replace("總計", "").strip()
            if not employee_name or employee_name in {"姓名", "投保級距", "非投保單位"}:
                continue

            # 依每個人員區塊的表頭，找出目標欄位位置
            header_window = row[start : min(start + 14, n_cols)]
            idx_project = 1
            idx_amount = 2
            idx_tax = next((k for k, v in enumerate(header_window) if "所得稅" in v or "應扣稅額" in v), None)
            idx_health = next((k for k, v in enumerate(header_window) if "二代健保" in v), None)
            idx_net = next((k for k, v in enumerate(header_window) if "實領" in v), None)
            idx_withhold_tax = next((k for k, v in enumerate(header_window) if "應扣所得" in v or "應扣所得稅" in v), None)
            idx_withhold_health = next((k for k, v in enumerate(header_window) if "應扣二代" in v), None)

            for r in range(i + 1, n_rows):
                cur = [clean_text(v) for v in df.iloc[r].tolist()]
                first = cur[start] if start < n_cols else ""
                project = cur[start + idx_project] if start + idx_project < n_cols else ""
                amount = to_number(cur[start + idx_amount] if start + idx_amount < n_cols else 0)

                # 遇到總計列或下一段人員表頭，結束此人員區塊
                if "總計" in first:
                    break
                if first and project == "案場":
                    break

                if not project or project in {"案場", "-"}:
                    continue
                if amount <= 0:
                    continue

                tax = to_number(cur[start + idx_tax] if idx_tax is not None and start + idx_tax < n_cols else 0)
                health = to_number(cur[start + idx_health] if idx_health is not None and start + idx_health < n_cols else 0)
                net = to_number(cur[start + idx_net] if idx_net is not None and start + idx_net < n_cols else amount)
                withhold_tax = to_number(
                    cur[start + idx_withhold_tax] if idx_withhold_tax is not None and start + idx_withhold_tax < n_cols else 0
                )
                withhold_health = to_number(
                    cur[start + idx_withhold_health] if idx_withhold_health is not None and start + idx_withhold_health < n_cols else 0
                )

                records.append(
                    {
                        "sheet_name": sheet_name,
                        "employee_name": employee_name,
                        "company_name": sheet_name,
                        "project_name": project,
                        "roc_year": roc_year,
                        "salary": amount,
                        "bonus": tax,
                        "welfare": health,
                        "total_income": net,
                        "note": f"應扣所得:{withhold_tax}, 應扣二代:{withhold_health}",
                    }
                )
    return records


def parse_salary_ratio_workbook(file_bytes: bytes) -> List[dict]:
    xls = pd.ExcelFile(BytesIO(file_bytes))
    target_sheet = "全案_總表" if "全案_總表" in xls.sheet_names else xls.sheet_names[0]
    df = pd.read_excel(BytesIO(file_bytes), sheet_name=target_sheet, header=3)
    df.columns = [str(c).strip().replace("\n", "") for c in df.columns]

    col_company = next((c for c in df.columns if "公司" in c), None)
    col_project = next((c for c in df.columns if "案場" in c or c == "案名"), None)
    col_total_sales = next((c for c in df.columns if "全案總銷" in c), None)
    col_booking = next((c for c in df.columns if c == "記帳"), None)
    col_contract = next((c for c in df.columns if "簽約金額" in c and "%" not in c), None)
    col_net_request = next((c for c in df.columns if "請款淨額" in c), None)
    col_revenue = next((c for c in df.columns if c == "營業收入"), None)
    col_sales_pct = next((c for c in df.columns if c == "總銷1%"), None)
    col_contract_pct = next((c for c in df.columns if "簽約金額1%" in c), None)
    col_salary = next((c for c in df.columns if "薪資" in c and "含" in c), None)
    if not col_salary:
        col_salary = next((c for c in df.columns if c == "薪資" or "薪資" in c), None)
    col_bonus = next((c for c in df.columns if "獎金" in c), None)
    col_welfare = next((c for c in df.columns if "福利" in c), None)
    col_total = next((c for c in df.columns if c == "總計"), None)

    if not col_company or not col_project:
        raise ValueError("無法辨識『全案_總表』的公司/案場欄位。")

    records: List[dict] = []
    for _, row in df.iterrows():
        company = str(row.get(col_company, "")).strip()
        project = str(row.get(col_project, "")).strip()
        if not company or company in {"nan", "合計"}:
            continue
        records.append(
            {
                "sheet_name": target_sheet,
                "employee_name": None,
                "company_name": company,
                "project_name": project,
                "roc_year": None,
                "salary": to_number(row.get(col_salary)),
                "bonus": to_number(row.get(col_bonus)),
                "welfare": to_number(row.get(col_welfare)),
                "total_income": to_number(row.get(col_total)),
                "note": append_note_parts(
                    [
                        "來源: 薪資占比檔",
                        f"全案總銷:{to_number(row.get(col_total_sales))}",
                        f"記帳:{to_number(row.get(col_booking))}",
                        f"簽約金額:{to_number(row.get(col_contract))}",
                        f"請款淨額:{to_number(row.get(col_net_request))}",
                        f"營業收入:{to_number(row.get(col_revenue))}",
                        f"總銷1%:{to_number(row.get(col_sales_pct))}",
                        f"簽約金額1%:{to_number(row.get(col_contract_pct))}",
                    ]
                ),
            }
        )
    return records


def parse_bonus_stat_workbook(file_bytes: bytes) -> List[dict]:
    xls = pd.ExcelFile(BytesIO(file_bytes))
    target_sheet = "在職年統計" if "在職年統計" in xls.sheet_names else xls.sheet_names[0]
    df = pd.read_excel(BytesIO(file_bytes), sheet_name=target_sheet, header=0)
    df.columns = [str(c).strip() for c in df.columns]

    records: List[dict] = []
    for i in [0, 1]:
        suffix = "" if i == 0 else ".1"
        col_name = f"姓名{suffix}" if f"姓名{suffix}" in df.columns else ("姓名" if i == 0 else None)
        if not col_name:
            continue
        col_113 = f"113年{suffix}" if f"113年{suffix}" in df.columns else None
        col_114 = f"114年{suffix}" if f"114年{suffix}" in df.columns else None
        col_115 = f"115年{suffix}" if f"115年{suffix}" in df.columns else None
        col_sum = f"合計{suffix}" if f"合計{suffix}" in df.columns else None

        for _, row in df.iterrows():
            name = str(row.get(col_name, "")).strip()
            if not name or name in {"nan", "姓名"}:
                continue
            records.append(
                {
                    "sheet_name": target_sheet,
                    "employee_name": name,
                    "company_name": None,
                    "project_name": None,
                    "roc_year": 114,
                    "salary": 0.0,
                    "bonus": to_number(row.get(col_114)),
                    "welfare": 0.0,
                    "total_income": to_number(row.get(col_sum)),
                    "note": f"113:{to_number(row.get(col_113))}, 114:{to_number(row.get(col_114))}, 115:{to_number(row.get(col_115))}",
                }
            )
    return records


def parse_personal_income_workbook(file_bytes: bytes) -> List[dict]:
    xls = pd.ExcelFile(BytesIO(file_bytes))
    records: List[dict] = []

    for sheet in xls.sheet_names:
        # 跳過彙總/說明頁，主抓每人每案場明細頁
        if str(sheet).startswith("總計-") or str(sheet).startswith("薪資-") or str(sheet) == "扣繳":
            continue

        df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet, header=None)
        if df.empty:
            continue
        records.extend(parse_personal_income_detail_sheet(df, sheet))
    return records


st.set_page_config(page_title="薪資報表匯入管理系統", layout="wide")
init_db()

APP_VERSION = "20260524-37"

st.markdown(
    """
    <style>
    /* 避免表單窄欄把日期日曆的月份選單裁切 */
    div[data-baseweb="popover"] {
        z-index: 10000 !important;
    }
    div[data-baseweb="popover"] > div,
    div[data-baseweb="calendar"],
    ul[role="listbox"] {
        overflow: visible !important;
        max-height: none !important;
    }

    /* ---- 介面美化 ---- */
    h1 {
        font-weight: 800;
        letter-spacing: 0.02em;
        background: linear-gradient(90deg, #2563eb, #7c3aed);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
    }
    /* 主頁籤：膠囊樣式 */
    .stTabs [data-baseweb="tab-list"] {
        gap: 6px;
        border-bottom: 1px solid rgba(120, 144, 180, 0.25);
        padding-bottom: 0;
    }
    .stTabs [data-baseweb="tab"] {
        border-radius: 10px 10px 0 0;
        padding: 8px 20px;
        background: rgba(37, 99, 235, 0.07);
        font-weight: 600;
    }
    .stTabs [aria-selected="true"] {
        background: #2563eb !important;
        color: #ffffff !important;
    }
    .stTabs [aria-selected="true"] p {
        color: #ffffff !important;
    }
    /* 表單卡片化 */
    [data-testid="stForm"] {
        border: 1px solid rgba(120, 144, 180, 0.28);
        border-radius: 14px;
        padding: 20px 20px 10px;
        box-shadow: 0 2px 10px rgba(16, 24, 40, 0.06);
    }
    /* 按鈕圓角 */
    .stButton > button,
    .stDownloadButton > button,
    [data-testid="stFormSubmitButton"] > button {
        border-radius: 10px;
        font-weight: 600;
    }
    /* 展開器 */
    [data-testid="stExpander"] {
        border: 1px solid rgba(120, 144, 180, 0.25);
        border-radius: 12px;
        overflow: hidden;
    }
    /* 表格圓角與陰影（勿設 overflow:hidden，會裁掉全視窗按鈕） */
    [data-testid="stDataFrame"] {
        border-radius: 12px;
        box-shadow: 0 1px 6px rgba(16, 24, 40, 0.06);
    }
    [data-testid="stDataFrame"] [data-testid="stElementToolbar"] {
        opacity: 1 !important;
        visibility: visible !important;
    }
    /* 報表 HTML 表格：標題列與合計同色藍底 */
    .report-table-wrap {
        overflow: auto;
        border-radius: 12px;
        box-shadow: 0 1px 6px rgba(16, 24, 40, 0.06);
        border: 1px solid rgba(120, 144, 180, 0.25);
        margin-bottom: 0.75rem;
        background: transparent;
    }
    .report-table-wrap table {
        width: 100%;
        border-collapse: collapse;
        font-size: 0.92rem;
    }
    .report-table-wrap thead th {
        position: sticky;
        top: 0;
        z-index: 3;
        background-color: #1d4ed8 !important;
        color: #ffffff !important;
        font-weight: 700 !important;
    }
    .report-table-wrap tbody td {
        white-space: nowrap;
    }
    /* 側欄 */
    [data-testid="stSidebar"] {
        border-right: 1px solid rgba(120, 144, 180, 0.2);
    }
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("人事成本管理系統")
st.caption(f"依「人事成本系統.xlsx」範本：全案總表、人事成本、在職年統計、個人所得。（版本 {APP_VERSION}）")

record_count = count_payroll_records()
batch_count = count_import_batches()
with st.sidebar:
    st.markdown("### 資料庫狀態")
    m1, m2 = st.columns(2)
    m1.metric("紀錄", f"{record_count:,}")
    m2.metric("批次", f"{batch_count:,}")
    st.caption(f"路徑：{DB_PATH}")

local_ip = get_local_ip()
with st.expander("跨裝置開啟網站"):
    st.write("在同一個 Wi-Fi/區網下，請用以下方式啟動：")
    st.code("streamlit run app.py --server.address 0.0.0.0 --server.port 8501")
    st.write(f"其他裝置可開：`http://{local_ip}:8501`（請確認 Windows 防火牆允許 8501 連線）")

tab_import, tab_report, tab_manual, tab_query, tab_batches = st.tabs(
    ["匯入資料", "報表呈現", "手動新增", "資料查詢", "匯入紀錄"]
)

with tab_import:
    st.subheader("檔案匯入（人事成本明細）")
    st.caption(
        "上傳 Excel / CSV，欄位：年度、公司名、案名、姓名、日期、項目、金額、勞保、勞退、"
        "保費、金額、稅款、金額、獎項、次數、備註（與手動新增同一筆可同時填寫各區塊）。"
    )
    col_tpl, col_hint = st.columns([1, 2])
    with col_tpl:
        st.download_button(
            "下載空白範本",
            data=build_hr_import_template_bytes(),
            file_name="人事成本_檔案匯入範本.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_hr_import_template",
        )
    with col_hint:
        st.info("項目＝薪資/三節/獎金/員工福利；保費＝健保/二代；稅款（或稅務）＝所得稅/執行業務所得。")

    uploader_key = st.session_state.get("upload_hr_detail_key", 0)
    detail_file = st.file_uploader(
        "上傳匯入檔",
        type=["xlsx", "xls", "csv"],
        key=f"upload_hr_detail_{uploader_key}",
    )
    st.caption("上傳後按「確認匯入」才會寫入。舊的「總表分表匯入」請到「匯入紀錄」刪除。")
    if detail_file is not None:
        file_bytes = detail_file.getvalue()
        file_hash = hashlib.md5(file_bytes).hexdigest()
        if st.session_state.get("hr_detail_file_hash") != file_hash:
            try:
                detail_records, detail_preview = parse_hr_detail_workbook(file_bytes, detail_file.name)
                st.session_state["hr_detail_import_records"] = detail_records
                st.session_state["hr_detail_import_preview"] = detail_preview
                st.session_state["hr_detail_import_filename"] = detail_file.name
                st.session_state["hr_detail_file_hash"] = file_hash
            except Exception as exc:
                st.error(f"讀檔失敗：{exc}")
                for key in ("hr_detail_import_records", "hr_detail_import_preview", "hr_detail_import_filename", "hr_detail_file_hash"):
                    st.session_state.pop(key, None)

    preview_df = st.session_state.get("hr_detail_import_preview")
    detail_records = st.session_state.get("hr_detail_import_records", [])
    if preview_df is not None and not (isinstance(preview_df, pd.DataFrame) and preview_df.empty):
        st.markdown("### 檔案匯入資料（預覽，尚未寫入）")
        st.dataframe(preview_df, use_container_width=True, hide_index=True)
        st.caption(f"共 {len(detail_records)} 筆可匯入。請按下方「確認匯入」。")
    elif detail_file is not None and not detail_records:
        st.warning("檔案中沒有可匯入的資料列，請確認欄位與範本一致。")

    if detail_records and st.button("確認匯入", type="primary", key="confirm_hr_detail_import"):
        try:
            batch_name = st.session_state.get("hr_detail_import_filename", "hr_detail_import")
            total, batch_id = save_import_records("人事成本", batch_name, detail_records)
            st.success(f"匯入成功：批次 #{batch_id}，共 {total} 筆。")
            st.info(f"本次為批次 #{batch_id}。若報表空白，請確認「匯入紀錄」已刪除舊的「總表分表匯入」等資料。")
            for key in (
                "hr_detail_import_records",
                "hr_detail_import_preview",
                "hr_detail_import_filename",
                "hr_detail_file_hash",
            ):
                st.session_state.pop(key, None)
            st.session_state["upload_hr_detail_key"] = uploader_key + 1
            st.rerun()
        except Exception as exc:
            st.error(f"匯入失敗：{exc}")

    st.divider()
    st.subheader("人事成本系統.xlsx（整份範本）")
    st.caption("請使用範本檔（含全案總表、人事成本等分頁）；若含「檔案匯入資料」分頁也會一併匯入。")
    hr_file = st.file_uploader("上傳人事成本系統檔", type=["xlsx", "xls"], key="upload_hr_system")
    st.caption("匯入會累加在現有資料上；若要清空請到「匯入紀錄」手動操作。")
    if hr_file is not None and st.button("匯入人事成本系統", key="import_hr_system"):
        try:
            parsed = parse_hr_system_workbook(hr_file.getvalue())
            total = 0
            batch_ids: list[int] = []
            for source_type, records in parsed.items():
                if records:
                    count, batch_id = save_import_records(source_type, hr_file.name, records)
                    total += count
                    if batch_id:
                        batch_ids.append(batch_id)
            batch_hint = f"（批次 {', '.join(f'#{i}' for i in batch_ids)}）" if batch_ids else ""
            st.success(
                f"匯入成功，共 {total} 筆{batch_hint}"
                f"（全案總表 {len(parsed.get('全案總表', []))}、人事成本 {len(parsed.get('人事成本', []))}）。"
            )
            st.info("在職年統計、個人所得會依人事成本資料自動計算顯示。")
            preview = parsed.get("全案總表", []) + parsed.get("人事成本", [])
            if preview:
                st.dataframe(pd.DataFrame(preview).head(20), use_container_width=True, hide_index=True)
        except Exception as exc:
            st.error(f"匯入失敗：{exc}")

with tab_report:
    st.subheader("報表呈現")
    records = list_payroll_records(limit=100000)
    if not records:
        st.info("目前沒有資料，請先到「匯入資料」上傳「人事成本系統.xlsx」。")
    else:
        df_all = pd.DataFrame([dict(r) for r in records])
        hidden_count = int((~df_all["source_type"].map(is_report_visible_source)).sum())
        if hidden_count:
            st.warning(
                f"資料庫有 {hidden_count} 筆舊格式（如「總表分表匯入」）不會出現在報表。"
                "請到「匯入紀錄」按「刪除報表不顯示的舊資料」，或刪除批次 #7。"
            )
        fy = st.selectbox("篩選年度", ["全部", *YEAR_OPTIONS], key="report_filter_year")
        filter_year = None if fy == "全部" else int(fy)
        report_view = st.selectbox("選擇報表", ["全案總表", "人事成本", "在職年統計", "個人所得", "月份總計"], key="report_view")

        def pick_keyword_filter(df: pd.DataFrame, column: str, label: str, key: str) -> pd.DataFrame:
            if df.empty or column not in df.columns:
                return df
            keyword = st.text_input(label, placeholder="輸入關鍵字篩選，空白顯示全部", key=key)
            if keyword.strip():
                q = keyword.strip().lower()
                return df[df[column].astype(str).str.lower().str.contains(q, na=False, regex=False)]
            return df

        def pick_dropdown_filter(df: pd.DataFrame, column: str, label: str, key: str) -> pd.DataFrame:
            if df.empty or column not in df.columns:
                return df
            options = ["全部"] + sorted({str(v).strip() for v in df[column].dropna().astype(str) if str(v).strip()})
            choice = st.selectbox(label, options, key=key)
            if choice != "全部":
                return df[df[column].astype(str).str.strip() == choice]
            return df

        def show_report_table(
            df: pd.DataFrame,
            cols: list[str],
            numeric_cols: list[str],
            title: str,
            file_name: str,
            key: str,
            percent_cols: list[str] | None = None,
            ratio_mode: str = "site",
        ) -> None:
            percent_cols = percent_cols or []
            if df.empty:
                st.warning(f"尚無「{title}」資料。")
                return
            visible = st.multiselect(f"{title}顯示欄位", cols, default=cols, key=f"visible_{key}")
            shown = visible if visible else cols
            num_cols = [c for c in numeric_cols if c in shown]
            display_df = df[shown].copy()
            for c in num_cols:
                display_df[c] = pd.to_numeric(display_df[c], errors="coerce").fillna(0.0)
            sum_cols = [c for c in num_cols if c not in percent_cols]
            total_kwargs: dict = {c: float(display_df[c].sum()) for c in sum_cols}
            # 「總計」欄改為各分項／各月合計，避免把各列總計再加總造成膨脹
            if "總計" in total_kwargs:
                part_cols = [c for c in sum_cols if c != "總計"]
                if part_cols:
                    total_kwargs["總計"] = float(sum(total_kwargs[c] for c in part_cols))
            if "比例" in percent_cols and "人事成本" in display_df.columns:
                hr_sum = float(display_df["人事成本"].sum())
                if ratio_mode == "hq":
                    den = hq_revenue_base(
                        float(display_df["營收"].sum()) if "營收" in display_df.columns else 0.0,
                        float(display_df["營收(未進帳)"].sum()) if "營收(未進帳)" in display_df.columns else 0.0,
                    )
                    total_kwargs["比例"] = calc_hr_ratio(hr_sum, den) if den > 0 else 0.0
                elif "請款額1%" in display_df.columns:
                    req_sum = float(display_df["請款額1%"].sum())
                    total_kwargs["比例"] = calc_hr_ratio(hr_sum, req_sum) if req_sum > 0 else 0.0
            text_cols = [c for c in shown if c not in num_cols]
            if text_cols:
                total_kwargs[text_cols[0]] = "合計"
                for c in text_cols[1:]:
                    total_kwargs[c] = ""
            display_df = pd.concat([display_df, pd.DataFrame([total_kwargs])], ignore_index=True)
            table_height = min(1400, max(560, 100 + len(display_df) * 42))
            render_report_table(
                format_currency_df(display_df, num_cols, percent_cols=percent_cols),
                max_height=table_height,
            )
            st.download_button(
                f"匯出{title} Excel",
                data=to_excel_bytes(display_df[shown], title, [c for c in num_cols if c in shown], title),
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_{key}",
            )

        if report_view == "全案總表":
            case_df = build_case_total_frame(df_all, filter_year)
            site_df = case_df[case_df["案場"] != HEADQUARTERS_PROJECT].copy()
            site_case_cols = [c for c in CASE_TOTAL_COLS if c not in {"營收", "營收(未進帳)"}]
            show_report_table(
                site_df,
                site_case_cols,
                [c for c in site_case_cols if c not in {"年度", "公司名", "案場"}],
                "全案總表",
                "全案總表_匯出.xlsx",
                "case_total",
                percent_cols=["比例"],
            )
            st.caption(
                "案場比例 = 人事成本 ÷ 請款額1% × 100；"
                "合計列比例 = 人事成本合計 ÷ 請款額1%合計 × 100。"
            )
            st.markdown("#### 總公司")
            hq_df = case_df[case_df["案場"] == HEADQUARTERS_PROJECT].copy()
            hq_case_cols = [
                c for c in CASE_TOTAL_COLS
                if c not in {"總銷", "簽約金額", "銷售請款額", "請款額1%", "請款淨額"}
            ]
            show_report_table(
                hq_df,
                hq_case_cols,
                [c for c in hq_case_cols if c not in {"年度", "公司名", "案場"}],
                "全案總表（總公司）",
                "全案總表_總公司_匯出.xlsx",
                "case_total_hq",
                percent_cols=["比例"],
                ratio_mode="hq",
            )
            st.caption("總公司比例(%) = 人事成本 ÷ (營收 + 營收(未進帳)) × 100。")
        elif report_view == "人事成本":
            hr_df = build_hr_cost_frame(df_all, filter_year)
            site_hr_df = hr_df[hr_df["案場"] != HEADQUARTERS_PROJECT].copy()
            site_hr_cols = [c for c in HR_COST_COLS if c != "公司名"]
            if not site_hr_df.empty:
                site_hr_df = site_hr_df.groupby(["年度", "案場"], as_index=False).sum(numeric_only=True)
                cost_parts = [c for c in ["勞保", "勞退", "健保", "二代", "薪資", "三節", "獎金", "員工福利"] if c in site_hr_df.columns]
                if cost_parts:
                    site_hr_df["總計"] = site_hr_df[cost_parts].sum(axis=1)
            site_hr_df = pick_dropdown_filter(site_hr_df, "案場", "篩選案場", "hr_cost_site_filter")
            show_report_table(
                site_hr_df,
                site_hr_cols,
                [c for c in site_hr_cols if c not in {"年度", "案場"}],
                "人事成本",
                "人事成本_匯出.xlsx",
                "hr_cost",
            )
            st.markdown("#### 總公司")
            hq_hr_df = hr_df[hr_df["案場"] == HEADQUARTERS_PROJECT].copy()
            hq_hr_cols = [c for c in HR_COST_COLS if c != "獎金"]
            if not hq_hr_df.empty:
                cost_parts = [c for c in ["勞保", "勞退", "健保", "二代", "薪資", "三節", "員工福利"] if c in hq_hr_df.columns]
                if "總計" in hq_hr_df.columns and cost_parts:
                    hq_hr_df = hq_hr_df.copy()
                    hq_hr_df["總計"] = hq_hr_df[cost_parts].sum(axis=1)
            show_report_table(
                hq_hr_df,
                hq_hr_cols,
                [c for c in hq_hr_cols if c not in {"年度", "公司名", "案場"}],
                "人事成本（總公司）",
                "人事成本_總公司_匯出.xlsx",
                "hr_cost_hq",
            )
            st.caption("案場表依「年度 + 案場」彙總；總公司依「年度 + 公司名」彙總。")
        elif report_view == "在職年統計":
            yearly_df = build_yearly_stat_frame(df_all)
            yearly_df = pick_keyword_filter(yearly_df, "姓名", "搜尋姓名", "yearly_name_filter")
            show_report_table(
                yearly_df,
                YEARLY_STAT_COLS,
                [c for c in YEARLY_STAT_COLS if c != "姓名"],
                "在職年統計",
                "在職年統計_匯出.xlsx",
                "yearly_stat",
            )
            st.caption("資料條件：人事成本中的「薪資 + 獎金」。")
        elif report_view == "個人所得":
            income_df = build_personal_income_frame(df_all, filter_year)
            income_df = pick_keyword_filter(income_df, "姓名", "搜尋姓名", "income_name_filter")
            show_report_table(
                income_df,
                PERSONAL_INCOME_COLS,
                [c for c in PERSONAL_INCOME_COLS if c not in {"年度", "案場", "姓名"}],
                "個人所得",
                "個人所得_匯出.xlsx",
                "personal_income",
            )
            st.caption("依「年度 + 案場 + 姓名」加總；金額 = 薪資 + 三節 + 獎金 + 員工福利。")
        else:
            monthly_df = build_monthly_total_frame(df_all, filter_year)
            site_monthly_df = monthly_df[monthly_df["案場"] != HEADQUARTERS_PROJECT].copy()
            site_monthly_cols = [c for c in MONTHLY_TOTAL_COLS if c != "公司名"]
            if not site_monthly_df.empty:
                site_monthly_df = site_monthly_df.groupby(["年度", "案場", "項目"], as_index=False).sum(numeric_only=True)
                month_cols = [c for c in MONTHLY_TOTAL_COLS if c in site_monthly_df.columns and c.endswith("月")]
                if month_cols and "總計" in site_monthly_df.columns:
                    site_monthly_df["總計"] = site_monthly_df[month_cols].sum(axis=1)
            site_monthly_df = pick_dropdown_filter(site_monthly_df, "案場", "篩選案場", "monthly_site_filter")
            show_report_table(
                site_monthly_df,
                site_monthly_cols,
                [c for c in site_monthly_cols if c not in {"年度", "案場", "項目"}],
                "月份總計",
                "月份總計_匯出.xlsx",
                "monthly_total",
            )
            st.markdown("#### 總公司")
            hq_monthly_df = monthly_df[monthly_df["案場"] == HEADQUARTERS_PROJECT].copy()
            show_report_table(
                hq_monthly_df,
                MONTHLY_TOTAL_COLS,
                [c for c in MONTHLY_TOTAL_COLS if c not in {"年度", "公司名", "案場", "項目"}],
                "月份總計（總公司）",
                "月份總計_總公司_匯出.xlsx",
                "monthly_total_hq",
            )
            st.caption(
                "依「年度 + 公司名 + 案場 + 項目（薪資、獎金）」彙總；年度與全案總表相同（依匯入年度）。"
                "獎金含三節；每月依發薪日期歸月：1 月 = 2 月整月、2 月 = 3 月整月，以此類推。"
                "（全案總表人事成本另含勞健退等，與本表薪資/獎金加總可能不同。）"
            )

with tab_manual:
    st.subheader("手動新增資料")
    st.caption("規則同範本檔各分頁下方說明。請款額1% 會依銷售請款額自動計算（1%）。")

    manual_company_options = get_company_options()
    manual_project_options = get_project_options()
    manual_site_project_options = [p for p in manual_project_options if p != HEADQUARTERS_PROJECT]

    with st.expander("➕ 新增案場 / 公司選項"):
        a1, a2 = st.columns(2)
        with a1:
            new_company_name = st.text_input("新增公司名", key="new_company_option", placeholder="例如 新公司")
            if st.button("加入公司", key="add_company_option_btn"):
                if not new_company_name.strip():
                    st.error("請輸入公司名。")
                elif new_company_name.strip() in manual_company_options:
                    st.info("此公司已在選單中。")
                else:
                    add_custom_option("company", new_company_name.strip())
                    st.success(f"已加入公司「{new_company_name.strip()}」。")
                    st.rerun()
        with a2:
            new_project_name = st.text_input("新增案場", key="new_project_option", placeholder="例如 新案場")
            if st.button("加入案場", key="add_project_option_btn"):
                if not new_project_name.strip():
                    st.error("請輸入案場名。")
                elif new_project_name.strip() in manual_project_options:
                    st.info("此案場已在選單中。")
                else:
                    add_custom_option("project", new_project_name.strip())
                    st.success(f"已加入案場「{new_project_name.strip()}」。")
                    st.rerun()

        custom_companies = list_custom_options("company")
        custom_projects = list_custom_options("project")
        if custom_companies or custom_projects:
            st.divider()
            st.caption(
                "自訂公司：" + ("、".join(custom_companies) if custom_companies else "無")
                + "　｜　自訂案場：" + ("、".join(custom_projects) if custom_projects else "無")
            )
            remove_labels = [f"公司｜{n}" for n in custom_companies] + [f"案場｜{n}" for n in custom_projects]
            r1, r2 = st.columns([2, 1])
            with r1:
                remove_choice = st.selectbox("刪除自訂選項", remove_labels, key="remove_custom_option_select")
            with r2:
                st.write("")
                if st.button("刪除選項", key="remove_custom_option_btn"):
                    kind, name = remove_choice.split("｜", 1)
                    delete_custom_option("company" if kind == "公司" else "project", name)
                    st.success(f"已刪除{kind}「{name}」。")
                    st.rerun()

    mtab1, mtab2, mtab3 = st.tabs(["全案總表", "人事成本", "案場成本"])

    with mtab1:
        with st.form("manual_case_form", clear_on_submit=True):
            c1, c2, c3, c4 = st.columns(4)
            with c1:
                case_year = st.selectbox("年度", YEAR_OPTIONS, index=YEAR_OPTIONS.index("114"), key="m_case_year")
            with c2:
                case_company = st.selectbox("公司名", manual_company_options, key="m_case_company")
            with c3:
                case_project = st.selectbox("案名", manual_site_project_options, key="m_case_project")
            with c4:
                case_date = pick_manual_date("m_case_date")
            c5, c6, c7, c8 = st.columns(4)
            with c5:
                case_field = st.selectbox("項目", CASE_FIELD_OPTIONS, key="m_case_field")
            with c6:
                case_amount = st.number_input("金額", min_value=0.0, step=1000.0, format="%.0f", key="m_case_amount")
            with c7:
                if case_field in CASE_DELTA_FIELDS:
                    case_op = st.radio("操作", ["增加", "扣除（入帳）"], horizontal=True, key="m_case_op")
                else:
                    case_op = "增加"
            with c8:
                case_remark = st.text_input("備註", key="m_case_remark")
            if case_field == "總銷":
                st.caption("總銷採覆蓋：新金額會取代舊值，不會累加。")
            elif case_field in CASE_DELTA_FIELDS:
                st.caption("營收(未進帳)可累加；選「扣除」代表入帳後減少未進帳金額。")
            if case_field == "銷售請款額" and case_amount > 0:
                st.info(f"請款額1% 自動帶入：{calc_request_pct(case_amount):,.0f}")
            submit_case = st.form_submit_button("新增全案總表資料（案場）")
        if submit_case:
            signed_amount = float(case_amount)
            if case_field in CASE_DELTA_FIELDS and case_op.startswith("扣除"):
                signed_amount = -abs(signed_amount)
            if case_field in CASE_OVERWRITE_FIELDS:
                note_parts = [
                    f"date:{case_date.isoformat()}",
                    "mode:overwrite",
                    f"field:{case_field}",
                    f"{case_field}:{signed_amount}",
                ]
            elif case_field in CASE_DELTA_FIELDS:
                note_parts = [
                    f"date:{case_date.isoformat()}",
                    "mode:delta",
                    f"field:{case_field}",
                    f"{case_field}:{signed_amount}",
                ]
            else:
                note_parts = [f"date:{case_date.isoformat()}", f"field:{case_field}", f"{case_field}:{signed_amount}"]
            if case_field == "銷售請款額":
                note_parts.append(f"請款額1%:{calc_request_pct(abs(case_amount))}")
            if case_remark.strip():
                note_parts.append(case_remark.strip())
            save_import_records(
                "全案總表手動",
                "manual_case",
                [{
                    "sheet_name": "全案總表",
                    "employee_name": None,
                    "company_name": case_company,
                    "project_name": case_project,
                    "roc_year": int(case_year),
                    "salary": 0.0,
                    "bonus": 0.0,
                    "welfare": 0.0,
                    "total_income": 0.0,
                    "note": append_note_parts(note_parts),
                }],
            )
            st.success("已新增全案總表資料（案場）。")
            st.rerun()

        st.markdown("#### 總公司")
        st.caption("總公司比例依「營收 + 營收(未進帳)」計算，不使用請款額1%。")
        with st.form("manual_hq_case_form", clear_on_submit=True):
            h1, h2, h3, h4 = st.columns(4)
            with h1:
                hq_year = st.selectbox("年度", YEAR_OPTIONS, index=YEAR_OPTIONS.index("114"), key="m_hq_year")
            with h2:
                hq_company = st.selectbox("公司名", manual_company_options, key="m_hq_company")
            with h3:
                hq_remark = st.text_input("備註", key="m_hq_remark")
            with h4:
                hq_date = pick_manual_date("m_hq_date")
            h5, h6, h7 = st.columns(3)
            with h5:
                hq_field = st.selectbox("項目", HQ_CASE_FIELD_OPTIONS, key="m_hq_field")
            with h6:
                hq_amount = st.number_input("金額", min_value=0.0, step=1000.0, format="%.0f", key="m_hq_amount")
            with h7:
                if hq_field in CASE_DELTA_FIELDS:
                    hq_op = st.radio("操作", ["增加", "扣除（入帳）"], horizontal=True, key="m_hq_op")
                else:
                    hq_op = "增加"
            if hq_field in CASE_DELTA_FIELDS:
                st.caption("營收(未進帳)可累加；選「扣除」代表入帳後減少未進帳金額。")
            submit_hq = st.form_submit_button(f"新增全案總表資料（{HEADQUARTERS_PROJECT}）")
        if submit_hq:
            signed_amount = float(hq_amount)
            if hq_field in CASE_DELTA_FIELDS and hq_op.startswith("扣除"):
                signed_amount = -abs(signed_amount)
            if hq_field in CASE_DELTA_FIELDS:
                note_parts = [
                    f"date:{hq_date.isoformat()}",
                    "mode:delta",
                    f"field:{hq_field}",
                    f"{hq_field}:{signed_amount}",
                ]
            else:
                note_parts = [f"date:{hq_date.isoformat()}", f"field:{hq_field}", f"{hq_field}:{signed_amount}"]
            if hq_remark.strip():
                note_parts.append(hq_remark.strip())
            save_import_records(
                "全案總表手動",
                "manual_hq_case",
                [{
                    "sheet_name": "全案總表",
                    "employee_name": None,
                    "company_name": hq_company,
                    "project_name": HEADQUARTERS_PROJECT,
                    "roc_year": int(hq_year),
                    "salary": 0.0,
                    "bonus": 0.0,
                    "welfare": 0.0,
                    "total_income": 0.0,
                    "note": append_note_parts(note_parts),
                }],
            )
            st.success(f"已新增全案總表資料（{HEADQUARTERS_PROJECT}）。")
            st.rerun()

    with mtab2:
        with st.form("manual_hr_form", clear_on_submit=True):
            c1, c2, c3, c4, c5 = st.columns(5)
            with c1:
                hr_year = st.selectbox("年度", YEAR_OPTIONS, index=YEAR_OPTIONS.index("114"), key="m_hr_year")
            with c2:
                hr_company = st.selectbox("公司名", manual_company_options, key="m_hr_company")
            with c3:
                hr_name = st.text_input("姓名", key="m_hr_name")
            with c4:
                hr_project = st.selectbox("案場", manual_project_options, key="m_hr_project")
            with c5:
                hr_date = pick_manual_date("m_hr_date")

            st.markdown("**① 項目**")
            h1, h2, h3, h4 = st.columns(4)
            with h1:
                hr_item = st.selectbox("項目", HR_MANUAL_ITEM_OPTIONS, key="m_hr_item")
            with h2:
                hr_item_amount = st.number_input("金額", min_value=0.0, step=100.0, format="%.0f", key="m_hr_item_amount")
            with h3:
                hr_bonus_type = st.selectbox("獎項", [""] + BONUS_TYPE_OPTIONS, key="m_hr_bonus_type")
            with h4:
                hr_times = st.selectbox("次數", [str(i) for i in range(1, 21)], key="m_hr_times")

            st.markdown("**② 法扣**")
            d1, d2 = st.columns(2)
            with d1:
                hr_labor = st.number_input("勞保", min_value=0.0, step=100.0, format="%.0f", key="m_hr_labor")
            with d2:
                hr_pension = st.number_input("勞退", min_value=0.0, step=100.0, format="%.0f", key="m_hr_pension")

            st.markdown("**③ 保費**")
            i1, i2 = st.columns(2)
            with i1:
                hr_health = st.number_input("健保", min_value=0.0, step=100.0, format="%.0f", key="m_hr_health")
            with i2:
                hr_nhi2 = st.number_input("二代", min_value=0.0, step=100.0, format="%.0f", key="m_hr_nhi2")

            st.markdown("**④ 稅務**")
            t1, t2 = st.columns(2)
            with t1:
                hr_income_tax = st.number_input("所得稅", min_value=0.0, step=100.0, format="%.0f", key="m_hr_income_tax")
            with t2:
                hr_business = st.number_input("執行業務所得", min_value=0.0, step=100.0, format="%.0f", key="m_hr_business")

            hr_note_text = st.text_input("備註（可空白）", key="m_hr_note")
            submit_hr = st.form_submit_button("新增人事成本資料")
        if submit_hr:
            if not hr_name.strip():
                st.error("請輸入姓名。")
            else:
                item_amounts = {
                    "薪資": float(hr_item_amount) if hr_item == "薪資" else 0.0,
                    "三節": float(hr_item_amount) if hr_item == "三節" else 0.0,
                    "獎金": float(hr_item_amount) if hr_item == "獎金" else 0.0,
                    "員工福利": float(hr_item_amount) if hr_item == "員工福利" else 0.0,
                }
                note_parts = [
                    f"date:{hr_date.isoformat()}",
                    f"勞保:{float(hr_labor)}",
                    f"勞退:{float(hr_pension)}",
                    f"健保:{float(hr_health)}",
                    f"二代:{float(hr_nhi2)}",
                    f"所得稅:{float(hr_income_tax)}",
                    f"執行業務所得:{float(hr_business)}",
                    f"薪資:{item_amounts['薪資']}",
                    f"三節:{item_amounts['三節']}",
                    f"獎金:{item_amounts['獎金']}",
                    f"員工福利:{item_amounts['員工福利']}",
                    f"次數:{hr_times}",
                ]
                if hr_bonus_type.strip():
                    note_parts.append(f"獎項:{hr_bonus_type.strip()}")
                if hr_note_text.strip():
                    note_parts.append(hr_note_text.strip())
                total = (
                    sum(item_amounts.values())
                    + float(hr_labor)
                    + float(hr_pension)
                    + float(hr_health)
                    + float(hr_nhi2)
                )
                save_import_records(
                    "人事成本手動",
                    "manual_hr",
                    [{
                        "sheet_name": "人事成本",
                        "employee_name": hr_name.strip(),
                        "company_name": hr_company,
                        "project_name": hr_project,
                        "roc_year": int(hr_year),
                        "salary": item_amounts["薪資"],
                        "bonus": item_amounts["獎金"],
                        "welfare": item_amounts["員工福利"],
                        "total_income": total,
                        "note": append_note_parts(note_parts),
                    }],
                )
                st.success("已新增人事成本資料（在職年統計、個人所得會自動連動）。")
                st.rerun()

    with mtab3:
        st.caption("案場成本歸屬人事成本：金額會計入該案場的人事成本（員工福利欄）。")
        with st.form("manual_site_cost_form", clear_on_submit=True):
            s1, s2, s3, s4 = st.columns(4)
            with s1:
                sc_year = st.selectbox("年度", YEAR_OPTIONS, index=YEAR_OPTIONS.index("114"), key="m_sc_year")
            with s2:
                sc_company = st.selectbox("公司名", manual_company_options, key="m_sc_company")
            with s3:
                sc_project = st.selectbox("案場", manual_project_options, key="m_sc_project")
            with s4:
                sc_date = pick_manual_date("m_sc_date")
            s5, s6, s7 = st.columns([1, 1, 2])
            with s5:
                sc_item = st.selectbox("項目", SITE_COST_ITEM_OPTIONS, key="m_sc_item")
            with s6:
                sc_amount = st.number_input("金額", min_value=0.0, step=100.0, format="%.0f", key="m_sc_amount")
            with s7:
                sc_remark = st.text_input("備註（可空白）", key="m_sc_remark")
            submit_sc = st.form_submit_button("新增案場成本資料")
        if submit_sc:
            if sc_amount <= 0:
                st.error("請輸入金額。")
            else:
                note_parts = [
                    f"date:{sc_date.isoformat()}",
                    f"員工福利:{float(sc_amount)}",
                    f"案場成本:{sc_item}",
                ]
                if sc_remark.strip():
                    note_parts.append(sc_remark.strip())
                save_import_records(
                    "人事成本手動",
                    "manual_site_cost",
                    [{
                        "sheet_name": "人事成本",
                        "employee_name": None,
                        "company_name": sc_company,
                        "project_name": sc_project,
                        "roc_year": int(sc_year),
                        "salary": 0.0,
                        "bonus": 0.0,
                        "welfare": float(sc_amount),
                        "total_income": float(sc_amount),
                        "note": append_note_parts(note_parts),
                    }],
                )
                st.success("已新增案場成本資料（計入人事成本的員工福利）。")
                st.rerun()

with tab_query:
    st.subheader("已匯入資料查詢")
    st.caption(
        "逐筆顯示，不合併。"
        "「資料ID」= 單筆編號（刪除請用這個）；「批次ID」= 同一批匯入的編號（刪整批請到匯入紀錄）。"
    )
    q1, q2, q3 = st.columns([2, 1, 1])
    with q1:
        keyword = st.text_input("關鍵字（姓名/公司/案場）", placeholder="輸入關鍵字")
    with q2:
        source_type = st.selectbox(
            "來源",
            ["全部", "全案總表", "全案總表手動", "人事成本", "人事成本手動"],
        )
    with q3:
        year_input = st.text_input("年度(民國)", placeholder="例如 114")
    roc_year = int(year_input) if year_input.strip().isdigit() else None

    rows = list_payroll_records(keyword=keyword, source_type=source_type, roc_year=roc_year, limit=100000)
    if rows:
        df = pd.DataFrame([dict(r) for r in rows])
        hidden = df[~df["source_type"].map(is_report_visible_source)]
        if not hidden.empty:
            st.warning(
                f"有 {len(hidden)} 筆來源為「{', '.join(sorted(hidden['source_type'].unique()))}」—"
                "這類舊資料不會出現在報表，可到「匯入紀錄」刪除。"
            )
        base_cols = [
            "id",
            "batch_id",
            "source_type",
            "sheet_name",
            "employee_name",
            "company_name",
            "project_name",
            "roc_year",
            "salary",
            "bonus",
            "welfare",
            "total_income",
        ]
        editable_df = df[base_cols].copy()
        editable_df = editable_df.rename(
            columns={
                "id": "資料ID",
                "batch_id": "批次ID",
                "source_type": "來源",
                "sheet_name": "分頁",
                "employee_name": "姓名",
                "company_name": "公司名",
                "project_name": "案場",
                "roc_year": "年度",
                "salary": "薪資",
                "bonus": "獎金",
                "welfare": "員工福利",
                "total_income": "總計",
            }
        )
        editable_df["獎項"] = df["note"].map(lambda n: parse_note_value(n, "獎項"))
        editable_df["次數"] = df["note"].map(lambda n: parse_note_value(n, "次數") or "1")
        editable_df["備註"] = df["note"].map(parse_note_remark)
        note_by_id = df.set_index("id")["note"].to_dict()
        money_cols = ["薪資", "獎金", "員工福利", "總計"]
        query_height = min(900, max(360, 80 + len(editable_df) * 38))
        render_report_table(
            format_currency_df(editable_df, money_cols),
            max_height=query_height,
        )
        with st.expander("編輯資料（修改後請按儲存）", expanded=False):
            edited = st.data_editor(
                editable_df,
                use_container_width=True,
                hide_index=True,
                height=min(520, query_height),
                key="query_inline_editor",
            )
            if st.button("儲存目前編輯", key="save_inline_edit"):
                for _, row in edited.iterrows():
                    rid = int(row["資料ID"])
                    old_note = note_by_id.get(rid, "")
                    update_payroll_record(
                        rid,
                        {
                            "sheet_name": row["分頁"],
                            "employee_name": row["姓名"],
                            "company_name": row["公司名"],
                            "project_name": row["案場"],
                            "roc_year": None if pd.isna(row["年度"]) else int(row["年度"]),
                            "salary": float(row["薪資"] or 0),
                            "bonus": float(row["獎金"] or 0),
                            "welfare": float(row["員工福利"] or 0),
                            "total_income": float(row["總計"] or 0),
                            "note": rebuild_note_display_fields(
                                old_note,
                                row["獎項"],
                                row["次數"],
                                row["備註"],
                            ),
                        },
                    )
                st.success("已儲存修改。")
                st.rerun()
        delete_ids_text = st.text_input(
            "刪除資料ID（逗號分隔）",
            placeholder="請填「資料ID」，例如 12,15",
            key="query_delete_ids",
        )
        if st.button("刪除指定資料ID", key="delete_inline_ids"):
            ids = [i.strip() for i in delete_ids_text.split(",") if i.strip()]
            deleted = delete_payroll_records([int(i) for i in ids if i.isdigit()])
            st.success(f"已刪除 {deleted} 筆。")
            st.rerun()
    else:
        st.info("查無資料。")

with tab_batches:
    st.subheader("匯入批次紀錄")
    st.caption("可刪除單一批次（該次上傳的全部資料）；不影響其他批次或手動新增的資料。")

    st.markdown("#### 刪除舊資料")
    st.info(
        "「總表分表匯入」等舊格式只會出現在資料查詢，**不會進報表**。"
        "你的截圖 batch_id=7、來源=總表分表匯入 就是這種，請用下方按鈕刪除。"
    )
    confirm_hidden = st.checkbox("確認刪除報表不顯示的舊資料", key="confirm_delete_hidden")
    if st.button("刪除報表不顯示的舊資料", type="primary", disabled=not confirm_hidden, key="delete_hidden_btn"):
        deleted, stats = delete_non_report_data()
        if deleted:
            st.success(f"已刪除 {deleted} 筆（含總表分表匯入等）。")
            if stats:
                st.json(stats)
        else:
            st.info("沒有需要刪除的舊資料。")
        st.rerun()

    batches = list_batches()
    total_batches = count_import_batches()
    st.caption(
        f"目前共 **{total_batches}** 筆匯入批次（顯示最新 {len(batches)} 筆，**沒有只顯示本週**的限制）。"
        "若找不到更早的上傳，可能當時已刪除，或部署時資料庫曾被重置。"
    )
    source_stats = count_batches_by_source()
    if source_stats:
        with st.expander("依來源統計批次", expanded=False):
            st.dataframe(pd.DataFrame([dict(r) for r in source_stats]), use_container_width=True, hide_index=True)

    if batches:
        batch_df = pd.DataFrame([dict(r) for r in batches])
        f1, f2 = st.columns([2, 1])
        with f1:
            batch_search = st.text_input(
                "搜尋批次（編號 / 來源 / 檔名）",
                placeholder="例如 11、人事成本、xlsx；清空可看全部",
                key="batch_search",
            )
        with f2:
            source_choices = ["全部"] + sorted({str(b["source_type"] or "") for b in batches})
            batch_source = st.selectbox("來源篩選", source_choices, key="batch_source_filter")

        view_df = batch_df.copy()
        if batch_source != "全部":
            view_df = view_df[view_df["source_type"].astype(str) == batch_source]
        if batch_search.strip():
            q = batch_search.strip().lower()
            mask = view_df.astype(str).apply(
                lambda row: any(q in str(v).lower() for v in row),
                axis=1,
            )
            view_df = view_df[mask]
            if view_df.empty:
                st.info("查無符合的匯入批次。請清空搜尋關鍵字，或改選「全部」來源。")
        if not view_df.empty:
            st.dataframe(view_df, use_container_width=True, hide_index=True)
            st.caption(f"目前列表顯示 {len(view_df)} 筆。檔案匯入通常來源為「人事成本」，檔名為上傳檔名。")

        batch_options = {
            f"#{b['id']}｜{b['source_type']}｜{b['file_name']}｜{b['row_count']}筆｜{b['imported_at']}": int(b["id"])
            for b in batches
            if (batch_source == "全部" or str(b["source_type"] or "") == batch_source)
            and (
                not batch_search.strip()
                or batch_search.strip().lower() in f"#{b['id']} {b['source_type']} {b['file_name']}".lower()
            )
        }
        if batch_options:
            selected_label = st.selectbox("選擇要刪除的批次", list(batch_options.keys()), key="delete_batch_select")
            confirm_batch = st.checkbox("確認刪除此批次", key="confirm_delete_batch")
            if st.button("刪除此批次", type="primary", disabled=not confirm_batch, key="delete_batch_btn"):
                bid = batch_options[selected_label]
                deleted_records, deleted_batches = delete_import_batch(bid)
                st.success(f"已刪除批次 #{bid}：{deleted_records} 筆資料。")
                st.rerun()
        elif batch_search.strip() or batch_source != "全部":
            st.info("查無符合的匯入批次可刪除。")
    else:
        st.info("尚無匯入紀錄。")

    st.markdown("### 手動清空全部資料")
    st.warning("僅在您要全部重來時使用；平常匯入與手動新增不會刪除資料。請先下載備份。")
    confirm_clear = st.checkbox("我了解並確認要清空全部資料", key="confirm_clear_all")
    if st.button("清空全部資料", type="primary", disabled=not confirm_clear, key="clear_all_data_btn"):
        deleted_records, deleted_batches = clear_all_data()
        st.success(f"已清空：刪除 {deleted_records} 筆紀錄、{deleted_batches} 筆匯入批次。")
        st.rerun()

    st.markdown("### 資料庫備份與還原")
    db_file = Path(DB_PATH)
    if db_file.exists():
        with db_file.open("rb") as f:
            st.download_button(
                "下載資料庫備份（.db）",
                data=f.read(),
                file_name=f"financial_reports_backup_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.db",
                mime="application/octet-stream",
                key="download_db_backup",
            )
    else:
        st.info("目前尚未建立資料庫檔案，先新增或匯入資料後再備份。")

    restore_file = st.file_uploader("上傳備份檔還原（.db）", type=["db"], key="restore_db_uploader")
    if restore_file is not None:
        st.warning("還原會覆蓋目前資料庫，請先下載備份。")
        if st.button("確認還原資料庫", key="confirm_restore_db"):
            try:
                db_file.parent.mkdir(parents=True, exist_ok=True)
                with db_file.open("wb") as f:
                    f.write(restore_file.getvalue())
                st.success("資料庫已還原，請重新整理頁面。")
                st.rerun()
            except Exception as exc:
                st.error(f"還原失敗：{exc}")
